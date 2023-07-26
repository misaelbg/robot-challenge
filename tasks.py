from RPA.Browser.Selenium import Selenium
from RPA.Robocorp.WorkItems import WorkItems
import time
import re
import os
import openpyxl
from openpyxl.utils import get_column_letter


class NYTimesScraper:
    def __init__(self, search_phrase, news_category, num_months):
        self.search_phrase = search_phrase
        self.news_category = news_category
        self.num_months = num_months
        self.browser_lib = Selenium()

    def open_the_website(self):
        self.browser_lib.open_available_browser("https://www.nytimes.com/")

    def perform_search(self):
        search_field = "css:input[type='search']"
        self.browser_lib.input_text(search_field, self.search_phrase)
        self.browser_lib.press_keys(search_field, "ENTER")

    def apply_filters(self):
        if self.news_category:
            category_xpath = f"//span[contains(text(), '{self.news_category}')]"
            self.browser_lib.click_element(category_xpath)
        time.sleep(2)

    def extract_data(self):
        articles = self.browser_lib.find_elements("css:li.css-ye6x8s")
        data = []
        for article in articles:
            title_element = article.find_element("css:h4")
            title = title_element.text
            description_element = article.find_element("css:p")
            description = description_element.text if description_element else ""
            date_element = article.find_element("css:time.css-1xti8j1")
            date = date_element.get_attribute("datetime")

            money_in_title = self.contains_money(title)
            money_in_description = self.contains_money(description)

            data.append((title, date, description, money_in_title or money_in_description))

        return data

    def contains_money(self, text):
        money_pattern = r"\$[\d,.]+|\d+\s*(dollars|USD)"
        return bool(re.search(money_pattern, text))

    def save_data_to_excel(self, data):
        excel_file = "output/news_data.xlsx"
        if os.path.exists(excel_file):
            os.remove(excel_file)

        wb = openpyxl.Workbook()
        ws = wb.active

        headers = ["Title", "Date", "Description", "Money in Title/Description"]
        for col_num, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = header

        for row_num, entry in enumerate(data, start=2):
            ws[f"A{row_num}"] = entry[0]  # Title
            ws[f"B{row_num}"] = entry[1]  # Date
            ws[f"C{row_num}"] = entry[2]  # Description
            ws[f"D{row_num}"] = entry[3]  # Money in Title/Description

        wb.save(excel_file)

    def download_images(self, data):
        image_folder = "output/images"
        os.makedirs(image_folder, exist_ok=True)

        for i, entry in enumerate(data, start=2):
            title = entry[0]
            image_file = f"{image_folder}/{i}.png"
            image_element = self.browser_lib.find_element("xpath://img[@alt='%s']" % title)
            image_element.screenshot(image_file)

    def run(self):
        try:
            self.open_the_website()
            self.perform_search()
            self.apply_filters()
            data = self.extract_data()
            self.save_data_to_excel(data)
            self.download_images(data)
        finally:
            self.browser_lib.close_all_browsers()

def main():
    work_items = WorkItems()
    work_items.get_input_work_item()
    search_phrase = work_items.get_work_item_variable("search_phrase")
    news_category = work_items.get_work_item_variable("category_or_section")
    num_months = work_items.get_work_item_variable("number_of_months")

    scraper = NYTimesScraper(search_phrase, news_category, num_months)
    scraper.run()

if __name__ == "__main__":
    main()
